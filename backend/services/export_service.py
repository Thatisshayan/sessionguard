"""
backend/services/export_service.py
------------------------------------
Generates real export artifacts using correct ReportLab 4.x API.

  PDF   → ReportLab 4.x  (full session + global reports with charts)
  Excel → openpyxl       (session detail + events, or global multi-sheet)

Maturity: Working Prototype
Future:   Evidence package builder (V7), branded letterhead (V9).
"""

import json
from datetime import datetime
from pathlib import Path

from database.db import get_connection
from engines.analysis_engine import (
    get_session_metrics, get_global_metrics,
    get_performance_by_game, get_net_result_over_time, get_rtp_distribution,
)
from engines.insights_engine import get_insights
from engines.alerts_engine import get_alerts

BASE_DIR    = Path(__file__).resolve().parent.parent.parent
EXPORTS_DIR = BASE_DIR / "storage" / "exports"
EXPORTS_DIR.mkdir(parents=True, exist_ok=True)


# ══════════════════════════════════════════════════════════════════════════════
# PDF EXPORT  (ReportLab 4.x)
# ══════════════════════════════════════════════════════════════════════════════

def generate_pdf(session_id: int | None = None) -> dict:
    """
    Generate a dark-themed PDF report.
    session_id=None → global summary.
    session_id=N    → single session deep-dive.
    """
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import mm
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT
        from reportlab.lib.colors import HexColor, white, black
        from reportlab.platypus import (
            SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
            HRFlowable, PageBreak,
        )
        # Correct ReportLab 4.x chart imports
        from reportlab.graphics.shapes import Drawing, Rect, String, Line
        from reportlab.graphics.charts.barcharts import VerticalBarChart
        from reportlab.graphics.charts.lineplots import LinePlot
        from reportlab.graphics.renderPDF import GraphicsFlowable
    except ImportError as e:
        return {"success": False, "file_path": "", "filename": "", "error": str(e)}

    ts       = datetime.now().strftime("%Y%m%d_%H%M%S")
    label    = f"session_{session_id}" if session_id else "global_summary"
    filename = f"sessionguard_report_{label}_{ts}.pdf"
    filepath = EXPORTS_DIR / filename

    # ── Colour palette matching frontend tokens ────────────────────────────────
    RL_BG    = HexColor("#0a0c10")
    RL_SURF  = HexColor("#111318")
    RL_ELEV  = HexColor("#1a1e26")
    RL_BORD  = HexColor("#242830")
    RL_TEXT  = HexColor("#e8eaf0")
    RL_MUTED = HexColor("#8892a4")
    RL_BLUE  = HexColor("#3b82f6")
    RL_GREEN = HexColor("#22c55e")
    RL_RED   = HexColor("#ef4444")
    RL_AMBER = HexColor("#f59e0b")

    doc = SimpleDocTemplate(
        str(filepath), pagesize=A4,
        leftMargin=18*mm, rightMargin=18*mm,
        topMargin=18*mm,  bottomMargin=18*mm,
    )

    # ── Styles ─────────────────────────────────────────────────────────────────
    def S(name, **kw):
        base = ParagraphStyle(name, fontName="Helvetica", fontSize=10,
                              textColor=RL_TEXT, leading=14)
        for k, v in kw.items():
            setattr(base, k, v)
        return base

    S_TITLE  = S("Title",  fontSize=22, textColor=RL_TEXT,  leading=28, spaceAfter=4,  fontName="Helvetica-Bold")
    S_H2     = S("H2",     fontSize=14, textColor=RL_TEXT,  leading=20, spaceBefore=10, spaceAfter=6, fontName="Helvetica-Bold")
    S_H3     = S("H3",     fontSize=11, textColor=RL_MUTED, leading=16, spaceBefore=6,  spaceAfter=4, fontName="Helvetica-Bold")
    S_BODY   = S("Body",   fontSize=10, textColor=RL_MUTED, leading=15, spaceAfter=4)
    S_MONO   = S("Mono",   fontSize=9,  textColor=RL_TEXT,  leading=13, fontName="Courier")
    S_CRIT   = S("Crit",   fontSize=10, textColor=RL_RED,   leading=15, spaceAfter=4)
    S_WARN   = S("Warn",   fontSize=10, textColor=RL_AMBER, leading=15, spaceAfter=4)
    S_INFO   = S("Info",   fontSize=10, textColor=RL_BLUE,  leading=15, spaceAfter=4)
    S_GOOD   = S("Good",   fontSize=10, textColor=RL_GREEN, leading=15, spaceAfter=4)

    def sev_style(sev):
        return {"critical": S_CRIT, "warning": S_WARN, "info": S_INFO}.get(sev, S_BODY)

    def rule():
        return HRFlowable(width="100%", thickness=1, color=RL_BORD,
                          spaceAfter=8, spaceBefore=8)

    # ── Logo / header block ────────────────────────────────────────────────────
    def header_block(subtitle: str) -> Table:
        now  = datetime.now().strftime("%Y-%m-%d  %H:%M")
        logo = Paragraph(
            "<font color='#3b82f6'>Session</font><font color='#e8eaf0'>Guard</font>",
            S("Logo", fontSize=20, fontName="Helvetica-Bold", textColor=RL_TEXT)
        )
        right = Paragraph(f"<font color='#8892a4'>{now}</font>",
                          S("R", fontSize=9, textColor=RL_MUTED, alignment=TA_RIGHT))
        t = Table([[logo, right]], colWidths=["65%", "35%"])
        t.setStyle(TableStyle([
            ("BACKGROUND",   (0,0), (-1,-1), RL_SURF),
            ("BOX",          (0,0), (-1,-1), 1, RL_BORD),
            ("TOPPADDING",   (0,0), (-1,-1), 14),
            ("BOTTOMPADDING",(0,0), (-1,-1), 14),
            ("LEFTPADDING",  (0,0), (-1,-1), 16),
            ("RIGHTPADDING", (0,0), (-1,-1), 16),
            ("VALIGN",       (0,0), (-1,-1), "MIDDLE"),
        ]))
        return t

    # ── KPI strip ──────────────────────────────────────────────────────────────
    def kpi_strip(kpis: list[tuple]) -> Table:
        """kpis = [(label, value, colour_hex), ...]"""
        labels = [Paragraph(k[0], S("KL", fontSize=7, textColor=RL_MUTED,
                                     fontName="Helvetica-Bold")) for k in kpis]
        values = [Paragraph(k[1], S("KV", fontSize=17, fontName="Courier",
                                     textColor=HexColor(k[2]), leading=21)) for k in kpis]
        cw = [172 / len(kpis) * mm] * len(kpis)
        t  = Table([labels, values], colWidths=cw)
        t.setStyle(TableStyle([
            ("BACKGROUND",   (0,0), (-1,-1), RL_ELEV),
            ("BOX",          (0,0), (-1,-1), 1, RL_BORD),
            ("LINEBELOW",    (0,0), (-1,0),  1, RL_BORD),
            ("TOPPADDING",   (0,0), (-1,-1), 9),
            ("BOTTOMPADDING",(0,0), (-1,-1), 9),
            ("LEFTPADDING",  (0,0), (-1,-1), 10),
            ("RIGHTPADDING", (0,0), (-1,-1), 10),
            ("ALIGN",        (0,0), (-1,-1), "CENTER"),
        ]))
        return t

    # ── Bar chart drawing ──────────────────────────────────────────────────────
    def bar_chart(data_pairs: list[tuple], title: str,
                  width=460, height=130) -> GraphicsFlowable:
        """data_pairs = [(label, value), ...]"""
        d   = Drawing(width, height)
        bc  = VerticalBarChart()
        bc.x, bc.y      = 50, 20
        bc.width        = width - 60
        bc.height       = height - 40
        bc.data         = [[v for _, v in data_pairs]]
        bc.categoryAxis.categoryNames = [l for l, _ in data_pairs]
        bc.categoryAxis.labels.fontSize  = 8
        bc.categoryAxis.labels.fillColor = RL_MUTED
        bc.categoryAxis.strokeColor      = RL_BORD
        bc.valueAxis.labels.fontSize     = 8
        bc.valueAxis.labels.fillColor    = RL_MUTED
        bc.valueAxis.strokeColor         = RL_BORD
        bc.valueAxis.gridStrokeColor     = RL_BORD
        bc.valueAxis.gridStrokeDashArray = [2, 2]
        bc.bars[0].fillColor   = RL_BLUE
        bc.bars[0].strokeColor = RL_BLUE
        # Background rect
        bg = Rect(0, 0, width, height, fillColor=RL_SURF, strokeColor=RL_BORD, strokeWidth=1)
        d.add(bg)
        t_label = String(width/2, height - 10, title, fontSize=9,
                         fillColor=RL_MUTED, textAnchor="middle")
        d.add(t_label)
        d.add(bc)
        return GraphicsFlowable(d)

    # ── Detail table ───────────────────────────────────────────────────────────
    def detail_table(rows: list[list]) -> Table:
        col_w = [50*mm, 46*mm, 50*mm, 26*mm]
        t = Table(rows, colWidths=col_w)
        t.setStyle(TableStyle([
            ("BACKGROUND",   (0,0), (-1,-1), RL_ELEV),
            ("BOX",          (0,0), (-1,-1), 1, RL_BORD),
            ("INNERGRID",    (0,0), (-1,-1), 0.5, RL_BORD),
            ("TOPPADDING",   (0,0), (-1,-1), 8),
            ("BOTTOMPADDING",(0,0), (-1,-1), 8),
            ("LEFTPADDING",  (0,0), (-1,-1), 10),
            ("TEXTCOLOR",    (0,0), (-1,-1), RL_MUTED),
            ("TEXTCOLOR",    (1,0), (1,-1), RL_TEXT),
            ("TEXTCOLOR",    (3,0), (3,-1), RL_TEXT),
            ("FONTNAME",     (1,0), (1,-1), "Courier"),
            ("FONTNAME",     (3,0), (3,-1), "Courier"),
            ("FONTSIZE",     (0,0), (-1,-1), 9),
        ]))
        return t

    # ══════════════════════════════════════════════════════════════════════════
    # Build content
    # ══════════════════════════════════════════════════════════════════════════
    els = []

    if session_id:
        m = get_session_metrics(session_id)
        if not m:
            return {"success": False, "file_path": "", "filename": "", "error": "Session not found."}

        insights = get_insights(session_id=session_id)
        alerts   = get_alerts(session_id=session_id)
        conn     = get_connection()
        events   = conn.execute(
            "SELECT * FROM events WHERE session_id=? ORDER BY timestamp", (session_id,)
        ).fetchall()
        conn.close()

        net_col = "#22c55e" if m["net_result"] >= 0 else "#ef4444"
        rtp_col = "#ef4444" if m["rtp"] < 85 else "#f59e0b" if m["rtp"] < 96 else "#3b82f6"
        str_col = "#ef4444" if m["losing_streak"] > 15 else "#f59e0b" if m["losing_streak"] > 8 else "#e8eaf0"

        # ── Header ────────────────────────────────────────────────────────────
        els.append(header_block("Session Report"))
        els.append(Spacer(1, 10))
        els.append(Paragraph(m["name"], S_TITLE))
        els.append(Paragraph(f"{m['game_name']}  ·  {m['platform']}  ·  {m['date']}", S_BODY))
        els.append(Spacer(1, 8))

        # ── KPI strip ─────────────────────────────────────────────────────────
        sign = "+" if m["net_result"] >= 0 else ""
        els.append(kpi_strip([
            ("NET RESULT",    f"{sign}${m['net_result']:.2f}",          net_col),
            ("RTP",           f"{m['rtp']}%",                           rtp_col),
            ("SPINS",         str(m["spins"]),                          "#e8eaf0"),
            ("BIGGEST WIN",   f"${m['biggest_win']:.2f}",               "#22c55e"),
            ("LOSING STREAK", str(m["losing_streak"]),                  str_col),
            ("MAX DRAWDOWN",  f"${m.get('max_drawdown', 0):.2f}",       "#ef4444"),
        ]))
        els.append(Spacer(1, 14))

        # ── Session detail grid ────────────────────────────────────────────────
        els.append(Paragraph("Session Details", S_H2))
        els.append(detail_table([
            ["Start Balance",  f"${m['start_balance']:.2f}",  "End Balance",     f"${m['end_balance']:.2f}"],
            ["Total Wagered",  f"${m['total_bets']:.2f}",     "Total Returned",  f"${m['total_wins']:.2f}"],
            ["Duration",       f"{m['duration_minutes']} min", "Platform",        m["platform"]],
            ["Status",         m["status"],                   "Notes",           m["notes"] or "—"],
        ]))
        els.append(Spacer(1, 14))

        # ── Events bar chart (bet vs win by spin sample) ──────────────────────
        if events:
            sample = events[::max(1, len(events)//20)]  # up to 20 samples
            spin_pairs = [(str(i+1), round(e["win_amount"], 2)) for i, e in enumerate(sample)]
            els.append(Paragraph("Win Distribution (sampled spins)", S_H2))
            els.append(bar_chart(spin_pairs, "Win Amount per Sampled Spin"))
            els.append(Spacer(1, 14))

        # ── Insights ──────────────────────────────────────────────────────────
        if insights:
            els.append(rule())
            els.append(Paragraph("Intelligence Insights", S_H2))
            icons = {"critical": "▶ [CRITICAL]", "warning": "▶ [WARNING]", "info": "▶ [INFO]"}
            for ins in insights:
                els.append(Paragraph(
                    f"{icons.get(ins['severity'], '▶')}  {ins['text']}",
                    sev_style(ins["severity"])
                ))
            els.append(Spacer(1, 8))

        # ── Alerts ────────────────────────────────────────────────────────────
        if alerts:
            els.append(rule())
            els.append(Paragraph("Active Alerts", S_H2))
            icons = {"critical": "■ [CRITICAL]", "warning": "■ [WARNING]"}
            for al in alerts:
                els.append(Paragraph(
                    f"{icons.get(al['severity'], '■')}  {al['message']}",
                    sev_style(al["severity"])
                ))
            els.append(Spacer(1, 8))

        # ── Event log (last 30) ────────────────────────────────────────────────
        if events:
            els.append(PageBreak())
            els.append(Paragraph("Event Log (last 30 spins)", S_H2))
            hdr = [["#", "Timestamp", "Type", "Bet", "Win", "Balance", "Conf"]]
            rows = []
            for i, ev in enumerate(events[-30:], 1):
                rows.append([
                    str(i),
                    ev["timestamp"][:19] if ev["timestamp"] else "",
                    ev["event_type"],
                    f"${ev['bet_amount']:.2f}",
                    f"${ev['win_amount']:.2f}",
                    f"${ev['balance_after']:.2f}",
                    f"{ev['confidence_score']:.2f}",
                ])
            ev_table = Table(hdr + rows,
                             colWidths=[15*mm, 38*mm, 18*mm, 20*mm, 20*mm, 22*mm, 16*mm])
            ev_table.setStyle(TableStyle([
                ("BACKGROUND",   (0,0), (-1,0),  RL_ELEV),
                ("BACKGROUND",   (0,1), (-1,-1), RL_SURF),
                ("ROWBACKGROUNDS",(0,1),(-1,-1),  [RL_SURF, RL_ELEV]),
                ("TEXTCOLOR",    (0,0), (-1,0),  RL_MUTED),
                ("TEXTCOLOR",    (0,1), (-1,-1), RL_TEXT),
                ("BOX",          (0,0), (-1,-1), 1, RL_BORD),
                ("INNERGRID",    (0,0), (-1,-1), 0.5, RL_BORD),
                ("FONTSIZE",     (0,0), (-1,-1), 8),
                ("FONTNAME",     (0,0), (-1,0),  "Helvetica-Bold"),
                ("TOPPADDING",   (0,0), (-1,-1), 5),
                ("BOTTOMPADDING",(0,0), (-1,-1), 5),
                ("LEFTPADDING",  (0,0), (-1,-1), 6),
                ("ALIGN",        (0,0), (-1,-1), "CENTER"),
            ]))
            # Green wins
            for i, ev in enumerate(events[-30:], 1):
                if ev["win_amount"] > 0:
                    ev_table.setStyle(TableStyle([
                        ("TEXTCOLOR", (4, i), (4, i), RL_GREEN)
                    ]))
            els.append(ev_table)

    else:
        # ── Global report ──────────────────────────────────────────────────────
        gm  = get_global_metrics()
        bg  = get_performance_by_game()
        not_ = get_net_result_over_time()
        rtp_ = get_rtp_distribution()
        insights = get_insights(limit=15)
        alerts   = get_alerts()

        net_col = "#22c55e" if gm["total_net"] >= 0 else "#ef4444"

        els.append(header_block("Global Summary"))
        els.append(Spacer(1, 10))
        els.append(Paragraph("SessionGuard — Platform Overview", S_TITLE))
        els.append(Paragraph(
            f"All {gm['total_sessions']} sessions  ·  "
            f"Generated {datetime.now().strftime('%Y-%m-%d %H:%M')}",
            S_BODY
        ))
        els.append(Spacer(1, 8))

        sign = "+" if gm["total_net"] >= 0 else ""
        els.append(kpi_strip([
            ("TOTAL NET",     f"{sign}${gm['total_net']:.2f}",           net_col),
            ("AVG RTP",       f"{gm['avg_rtp']}%",                       "#3b82f6"),
            ("SESSIONS",      str(gm["total_sessions"]),                  "#e8eaf0"),
            ("TOTAL SPINS",   str(gm["total_spins"]),                     "#e8eaf0"),
            ("BEST WIN",      f"${gm['all_time_biggest_win']:.2f}",       "#22c55e"),
            ("FLAGGED",       str(gm["flagged_count"]),                   "#f59e0b"),
        ]))
        els.append(Spacer(1, 16))

        # RTP distribution chart
        if rtp_:
            els.append(Paragraph("RTP Distribution", S_H2))
            els.append(bar_chart(
                [(r["bucket"], r["count"]) for r in rtp_],
                "Sessions per RTP bucket"
            ))
            els.append(Spacer(1, 14))

        # Performance by game table
        if bg:
            els.append(rule())
            els.append(Paragraph("Performance by Game", S_H2))
            hdr  = [["Game", "Sessions", "Avg RTP", "Total Net", "Avg Net"]]
            rows = [[
                r["game_name"], str(r["sessions"]),
                f"{r['avg_rtp']}%",
                f"{'+'if r['total_net']>=0 else ''}${r['total_net']:.2f}",
                f"{'+'if r['avg_net']>=0 else ''}${r['avg_net']:.2f}",
            ] for r in bg]
            gt = Table(hdr + rows,
                       colWidths=[55*mm, 22*mm, 24*mm, 34*mm, 34*mm])
            gt.setStyle(TableStyle([
                ("BACKGROUND",   (0,0), (-1,0),  RL_ELEV),
                ("BACKGROUND",   (0,1), (-1,-1), RL_SURF),
                ("ROWBACKGROUNDS",(0,1),(-1,-1),  [RL_SURF, RL_ELEV]),
                ("TEXTCOLOR",    (0,0), (-1,0),  RL_MUTED),
                ("TEXTCOLOR",    (0,1), (-1,-1), RL_TEXT),
                ("BOX",          (0,0), (-1,-1), 1, RL_BORD),
                ("INNERGRID",    (0,0), (-1,-1), 0.5, RL_BORD),
                ("FONTSIZE",     (0,0), (-1,-1), 9),
                ("FONTNAME",     (0,0), (-1,0),  "Helvetica-Bold"),
                ("TOPPADDING",   (0,0), (-1,-1), 7),
                ("BOTTOMPADDING",(0,0), (-1,-1), 7),
                ("LEFTPADDING",  (0,0), (-1,-1), 10),
            ]))
            els.append(gt)
            els.append(Spacer(1, 14))

        # Critical findings
        critical = [i for i in insights if i["severity"] == "critical"]
        if critical:
            els.append(rule())
            els.append(Paragraph("Critical Findings", S_H2))
            for ins in critical[:8]:
                els.append(Paragraph(f"■ [CRITICAL]  {ins['text']}", S_CRIT))
                els.append(Paragraph(f"  → {ins['session_name']}", S_BODY))

        # Active unacknowledged alerts
        unacked = [a for a in alerts if not a["acknowledged"]]
        if unacked:
            els.append(rule())
            els.append(Paragraph(f"Unacknowledged Alerts ({len(unacked)})", S_H2))
            for al in unacked[:10]:
                els.append(Paragraph(
                    f"■ [{al['severity'].upper()}]  {al['message']}",
                    sev_style(al["severity"])
                ))

    doc.build(els)

    conn2 = get_connection()
    cur   = conn2.execute(
        "INSERT INTO exports (session_id, format, file_path) VALUES (?, 'pdf', ?)",
        (session_id, str(filepath))
    )
    export_id = cur.lastrowid
    conn2.commit()
    conn2.close()

    return {
        "success":   True,
        "export_id": export_id,
        "file_path": str(filepath),
        "filename":  filename,
        "error":     None,
    }


# ══════════════════════════════════════════════════════════════════════════════
# EXCEL EXPORT  (openpyxl)
# ══════════════════════════════════════════════════════════════════════════════

def generate_excel(session_id: int | None = None) -> dict:
    """Generate a styled Excel workbook."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError as e:
        return {"success": False, "file_path": "", "filename": "", "error": str(e)}

    ts       = datetime.now().strftime("%Y%m%d_%H%M%S")
    label    = f"session_{session_id}" if session_id else "global"
    filename = f"sessionguard_{label}_{ts}.xlsx"
    filepath = EXPORTS_DIR / filename

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ── Style helpers ──────────────────────────────────────────────────────────
    def fill(h):  return PatternFill("solid", fgColor=h)
    def fnt(h, bold=False, sz=10, mono=False):
        return Font(color=h, bold=bold, size=sz,
                    name="Courier New" if mono else "Calibri")
    def ctr(): return Alignment(horizontal="center", vertical="center", wrap_text=True)
    def lft(): return Alignment(horizontal="left",   vertical="center", wrap_text=True)

    BG_S = "111318"; BG_E = "1A1E26"; BG_B = "242830"
    C_W  = "E8EAF0"; C_M  = "8892A4"
    C_G  = "22C55E"; C_R  = "EF4444"; C_A  = "F59E0B"; C_BL = "3B82F6"

    def style_header(ws, row, cols):
        for c in range(1, cols + 1):
            cell = ws.cell(row, c)
            cell.fill = fill(BG_E); cell.font = fnt(C_M, bold=True, sz=9)
            cell.alignment = ctr()

    def style_row(ws, row, cols, alt=False):
        for c in range(1, cols + 1):
            cell = ws.cell(row, c)
            cell.fill = fill(BG_E if alt else BG_S)
            cell.font = fnt(C_W, sz=10)
            cell.alignment = lft()

    def autofit(ws, pad=4):
        for col in ws.columns:
            mx = max((len(str(c.value or "")) for c in col), default=8)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(mx + pad, 45)

    def net_font(cell, val):
        cell.font = Font(color=C_G if float(val or 0) >= 0 else C_R,
                         bold=True, size=10, name="Calibri")

    conn = get_connection()

    if session_id:
        m = get_session_metrics(session_id)
        if not m:
            return {"success": False, "file_path": "", "filename": "", "error": "Session not found."}

        insights = get_insights(session_id=session_id)
        alerts   = get_alerts(session_id=session_id)
        events   = conn.execute(
            "SELECT * FROM events WHERE session_id=? ORDER BY timestamp", (session_id,)
        ).fetchall()

        # ── Sheet 1: Summary ───────────────────────────────────────────────────
        ws = wb.create_sheet("Summary")
        ws.sheet_view.showGridLines = False
        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 24

        kpis = [
            ("Session Name",    m["name"]),
            ("Game",            m["game_name"]),
            ("Platform",        m["platform"]),
            ("Date",            m["date"]),
            ("Duration (min)",  m["duration_minutes"]),
            ("Start Balance",   f"${m['start_balance']:.2f}"),
            ("End Balance",     f"${m['end_balance']:.2f}"),
            ("Net Result",      f"{'+'if m['net_result']>=0 else ''}${m['net_result']:.2f}"),
            ("Total Wagered",   f"${m['total_bets']:.2f}"),
            ("Total Returned",  f"${m['total_wins']:.2f}"),
            ("RTP",             f"{m['rtp']}%"),
            ("Spins",           m["spins"]),
            ("Biggest Win",     f"${m['biggest_win']:.2f}"),
            ("Losing Streak",   m["losing_streak"]),
            ("Max Drawdown",    f"${m.get('max_drawdown', 0):.2f}"),
            ("Status",          m["status"]),
        ]
        for i, (lbl, val) in enumerate(kpis, 1):
            lc = ws.cell(i, 1, lbl)
            lc.fill = fill(BG_E); lc.font = fnt(C_M, bold=True); lc.alignment = lft()
            vc = ws.cell(i, 2, str(val))
            vc.fill = fill(BG_S); vc.font = fnt(C_W, mono=True); vc.alignment = lft()
            if lbl == "Net Result":
                net_font(vc, m["net_result"])

        # ── Sheet 2: Events ────────────────────────────────────────────────────
        if events:
            we = wb.create_sheet("Events")
            we.sheet_view.showGridLines = False
            hdrs = ["#", "Timestamp", "Type", "Bet ($)", "Win ($)", "Balance ($)", "Confidence", "Source"]
            for c, h in enumerate(hdrs, 1):
                we.cell(1, c, h)
            style_header(we, 1, len(hdrs))

            for r, ev in enumerate(events, 2):
                row_vals = [
                    r - 1,
                    ev["timestamp"][:19] if ev["timestamp"] else "",
                    ev["event_type"],
                    round(float(ev["bet_amount"]), 2),
                    round(float(ev["win_amount"]), 2),
                    round(float(ev["balance_after"]), 2),
                    round(float(ev["confidence_score"]), 2),
                    ev["source"],
                ]
                for c, v in enumerate(row_vals, 1):
                    we.cell(r, c, v)
                style_row(we, r, len(hdrs), r % 2 == 0)
                if float(ev["win_amount"]) > 0:
                    we.cell(r, 5).font = fnt(C_G, bold=True)
                if float(ev["confidence_score"]) < 0.80:
                    we.cell(r, 7).font = fnt(C_A, bold=True)
            autofit(we)

        # ── Sheet 3: Insights ──────────────────────────────────────────────────
        if insights:
            wi = wb.create_sheet("Insights")
            wi.sheet_view.showGridLines = False
            for c, h in enumerate(["Severity", "Category", "Text", "Created"], 1):
                wi.cell(1, c, h)
            style_header(wi, 1, 4)
            for r, ins in enumerate(insights, 2):
                wi.cell(r, 1, ins["severity"])
                wi.cell(r, 2, ins["category"])
                wi.cell(r, 3, ins["text"])
                wi.cell(r, 4, ins["created_at"][:16])
                style_row(wi, r, 4, r % 2 == 0)
                cmap = {"critical": C_R, "warning": C_A, "info": C_BL}
                wi.cell(r, 1).font = fnt(cmap.get(ins["severity"], C_W), bold=True)
            autofit(wi)

    else:
        # ── Global workbook ────────────────────────────────────────────────────
        gm  = get_global_metrics()
        bg  = get_performance_by_game()
        not_ = get_net_result_over_time()

        # Sheet 1: Summary KPIs
        ws = wb.create_sheet("Summary")
        ws.sheet_view.showGridLines = False
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 24

        kpis = [
            ("Total Sessions",       gm["total_sessions"]),
            ("Total Net Result",     f"{'+'if gm['total_net']>=0 else ''}${gm['total_net']:.2f}"),
            ("Average RTP",          f"{gm['avg_rtp']}%"),
            ("Average Net/Session",  f"{'+'if gm['avg_net']>=0 else ''}${gm['avg_net']:.2f}"),
            ("Total Wagered",        f"${gm['total_wagered']:.2f}"),
            ("Total Returned",       f"${gm['total_returned']:.2f}"),
            ("Total Spins",          gm["total_spins"]),
            ("All-Time Biggest Win", f"${gm['all_time_biggest_win']:.2f}"),
            ("Worst Losing Streak",  gm["worst_streak"]),
            ("Flagged Sessions",     gm["flagged_count"]),
            ("Generated",            datetime.now().strftime("%Y-%m-%d %H:%M")),
        ]
        for i, (lbl, val) in enumerate(kpis, 1):
            lc = ws.cell(i, 1, lbl)
            lc.fill = fill(BG_E); lc.font = fnt(C_M, bold=True); lc.alignment = lft()
            vc = ws.cell(i, 2, str(val))
            vc.fill = fill(BG_S); vc.font = fnt(C_W, mono=True); vc.alignment = lft()

        # Sheet 2: All Sessions
        sessions = conn.execute("SELECT * FROM sessions ORDER BY date DESC").fetchall()
        wss = wb.create_sheet("All Sessions")
        wss.sheet_view.showGridLines = False
        hdrs = ["ID", "Name", "Game", "Platform", "Date",
                "Net Result", "RTP %", "Spins", "Total Bets",
                "Total Wins", "Biggest Win", "Streak", "Status"]
        for c, h in enumerate(hdrs, 1):
            wss.cell(1, c, h)
        style_header(wss, 1, len(hdrs))

        for r, s in enumerate(sessions, 2):
            vals = [
                s["id"], s["name"], s["game_name"], s["platform"], s["date"],
                round(float(s["net_result"]), 2), f"{s['rtp']}%", s["spins"],
                round(float(s["total_bets"]), 2), round(float(s["total_wins"]), 2),
                round(float(s["biggest_win"]), 2), s["losing_streak"], s["status"],
            ]
            for c, v in enumerate(vals, 1):
                wss.cell(r, c, v)
            style_row(wss, r, len(hdrs), r % 2 == 0)
            net_font(wss.cell(r, 6), s["net_result"])
        autofit(wss)

        # Sheet 3: By Game
        if bg:
            wg = wb.create_sheet("By Game")
            wg.sheet_view.showGridLines = False
            gh = ["Game", "Sessions", "Avg RTP %", "Total Net", "Avg Net/Session"]
            for c, h in enumerate(gh, 1): wg.cell(1, c, h)
            style_header(wg, 1, len(gh))
            for r, row in enumerate(bg, 2):
                vals = [row["game_name"], row["sessions"],
                        f"{row['avg_rtp']}%",
                        round(row["total_net"], 2), round(row["avg_net"], 2)]
                for c, v in enumerate(vals, 1): wg.cell(r, c, v)
                style_row(wg, r, len(gh), r % 2 == 0)
                net_font(wg.cell(r, 4), row["total_net"])
                net_font(wg.cell(r, 5), row["avg_net"])
            autofit(wg)

        # Sheet 4: Net Over Time
        if not_:
            wt = wb.create_sheet("Net Over Time")
            wt.sheet_view.showGridLines = False
            for c, h in enumerate(["Date", "Daily Net", "Cumulative Net"], 1):
                wt.cell(1, c, h)
            style_header(wt, 1, 3)
            for r, row in enumerate(not_, 2):
                wt.cell(r, 1, row["date"])
                wt.cell(r, 2, round(row["daily_net"], 2))
                wt.cell(r, 3, round(row["cumulative_net"], 2))
                style_row(wt, r, 3, r % 2 == 0)
                net_font(wt.cell(r, 2), row["daily_net"])
                net_font(wt.cell(r, 3), row["cumulative_net"])
            autofit(wt)

    conn.close()
    wb.save(str(filepath))

    reg_conn = get_connection()
    cur = reg_conn.execute(
        "INSERT INTO exports (session_id, format, file_path) VALUES (?, 'excel', ?)",
        (session_id, str(filepath))
    )
    export_id = cur.lastrowid
    reg_conn.commit()
    reg_conn.close()

    return {
        "success":   True,
        "export_id": export_id,
        "file_path": str(filepath),
        "filename":  filename,
        "error":     None,
    }
